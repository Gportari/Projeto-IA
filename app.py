from flask import Flask, render_template, request, redirect, jsonify, send_from_directory
import threading
import time
from model_registry import ModelRegistry
import os
import json
import numpy as np
from sklearn.model_selection import train_test_split
from werkzeug.utils import secure_filename
# Import de funções de treino será feito de forma lazy dentro de rotinas específicas

app = Flask(__name__)
registry = ModelRegistry()

# Background training job registry
TRAIN_JOBS = {}
TRAIN_JOBS_LOCK = threading.Lock()

# Callback de progresso será definido localmente no início do fluxo de treino

def _create_model_instance(model_type, model_config, training_params):
    # Factory to get model instance for callback-driven training
    if model_type == 'logistic':
        from trainingModels.LogisticRegression import create_model_from_config as _create
    elif model_type == 'knn':
        from trainingModels.KNN import create_model_from_config as _create
    elif model_type == 'svm':
        from trainingModels.SVM import create_model_from_config as _create
    elif model_type == 'tree':
        from trainingModels.DecisionTree import create_model_from_config as _create
    elif model_type == 'rf':
        from trainingModels.RandomForest import create_model_from_config as _create
    else:
        raise ValueError(f"Unsupported model type: {model_type}")
    return _create(model_config, training_params)

def _load_config_by_filename(filename):
    model_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models_json')
    config_path = os.path.join(model_dir, filename)
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def _start_training_thread(job_id, model_type, config_file, dataset, training_params):
    def _runner():
        try:
            features = np.array(dataset['features'])
            labels = np.array(dataset['labels'])
            test_size = float(training_params.get('test_split', 0.2))
            X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=test_size, random_state=42)

            # Create model
            model_config = _load_config_by_filename(config_file)
            model = _create_model_instance(model_type, model_config, training_params)
            # Safely parse epochs and expected total
            raw_ep = training_params.get('epochs', 50)
            try:
                epochs = int(raw_ep)
                if epochs <= 0:
                    epochs = 50
            except Exception:
                epochs = 50

            # Estimate total epochs for progress (account for RF multiple estimators)
            def _safe_int(v, d):
                try:
                    if v is None:
                        return d
                    if isinstance(v, str) and v.strip().lower() in ('none', 'null', ''):
                        return d
                    iv = int(v)
                    return iv if iv > 0 else d
                except Exception:
                    return d

            total_epochs_est = epochs
            if model_type == 'rf':
                n_est = _safe_int(model_config.get('rf_n_estimators'), 10)
                total_epochs_est = epochs * n_est

            # Importar Keras de forma lazy e definir callback local para evitar
            # que a aplicação quebre ao iniciar sem TensorFlow instalado.
            from tensorflow import keras

            class EpochProgressCallback(keras.callbacks.Callback):
                def __init__(self, job_id, total_epochs):
                    super().__init__()
                    self.job_id = job_id
                    self.total_epochs = total_epochs
                    self.start_time = time.time()
                    self.last_epoch_time = self.start_time
                    self.completed_epochs_so_far = 0
                    self._last_epoch_in_run = -1

                def on_train_begin(self, logs=None):
                    self._last_epoch_in_run = -1

                def on_epoch_end(self, epoch, logs=None):
                    self._last_epoch_in_run = epoch
                    global_epoch = self.completed_epochs_so_far + (epoch + 1)
                    now = time.time()
                    elapsed = now - self.start_time
                    avg_epoch_time = elapsed / max(1, global_epoch)
                    remaining_epochs = max(0, self.total_epochs - global_epoch)
                    eta_seconds = max(0, int(avg_epoch_time * remaining_epochs))
                    with TRAIN_JOBS_LOCK:
                        job = TRAIN_JOBS.get(self.job_id, {})
                        job.update({
                            'epoch': global_epoch,
                            'total_epochs': self.total_epochs,
                            'loss': float(logs.get('loss', 0.0)) if logs else 0.0,
                            'accuracy': float(logs.get('accuracy', 0.0)) if logs else 0.0,
                            'val_loss': float(logs.get('val_loss', 0.0)) if logs and 'val_loss' in logs else None,
                            'val_accuracy': float(logs.get('val_accuracy', 0.0)) if logs and 'val_accuracy' in logs else None,
                            'eta_seconds': eta_seconds,
                            'status': 'training'
                        })
                        if job.get('cancel_requested'):
                            job['status'] = 'cancelling'
                            self.model.stop_training = True
                        TRAIN_JOBS[self.job_id] = job

                def on_train_end(self, logs=None):
                    if self._last_epoch_in_run is not None and self._last_epoch_in_run >= 0:
                        self.completed_epochs_so_far += (self._last_epoch_in_run + 1)

            callbacks = [EpochProgressCallback(job_id, total_epochs_est)]

            # Train with callbacks
            model.fit(X_train, y_train, validation_split=float(training_params.get('validation_split', 0.2)), callbacks=callbacks)

            # Evaluate unless canceled
            with TRAIN_JOBS_LOCK:
                cancelled = TRAIN_JOBS.get(job_id, {}).get('cancel_requested')
            if cancelled:
                with TRAIN_JOBS_LOCK:
                    TRAIN_JOBS[job_id].update({
                        'status': 'canceled',
                        'model_type': model_type,
                        'config_file': config_file,
                        'model_ref': model
                    })
            else:
                metrics = model.evaluate(X_test, y_test)
                results = {
                    'name': model_config.get('nome_modelo') or model_config.get('name', model_type),
                    **metrics
                }
                with TRAIN_JOBS_LOCK:
                    TRAIN_JOBS[job_id].update({
                        'status': 'completed',
                        'results': results,
                        'model_type': model_type,
                        'config_file': config_file,
                        'model_ref': model
                    })
        except Exception as e:
            with TRAIN_JOBS_LOCK:
                TRAIN_JOBS[job_id].update({
                    'status': 'error',
                    'error': str(e)
                })

    t = threading.Thread(target=_runner, daemon=True)
    t.start()
    return t

# Garantir que o diretório models_json exista
models_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models_json')
if not os.path.exists(models_dir):
    os.makedirs(models_dir)

# Diretório para armazenar datasets
datasets_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'datasets')
if not os.path.exists(datasets_dir):
    os.makedirs(datasets_dir)

@app.route('/')
def home():
    # Página inicial passa a ser Models
    return render_template('Models.html', model_list=registry.listar_modelos())

@app.route('/Models.html', methods=['GET'])
def models():
    # Carregar modelos dos arquivos JSON
    models_list = registry.listar_modelos()
    return render_template('Models.html', model_list=models_list)

@app.route('/ModelComparison.html')
def model_comparison():
    # Carregar modelos dos arquivos JSON para a página de comparação
    models_list = registry.listar_modelos()
    return render_template('ModelComparison.html', model_list=models_list)

@app.route('/Dataset.html')
def dataset_page():
    # Página de visualização e gerenciamento do dataset
    return render_template('Dataset.html')

@app.route('/DatasetEditor.html', methods=['GET'])
def dataset_editor():
    # Página de edição do dataset (renomear e visualizar metadados)
    filename = (request.args.get('file') or '').strip()
    if not filename:
        return render_template('DatasetEditor.html', error='Arquivo não especificado.', filename=None, size_h=None, rows=None)
    safe_name = os.path.basename(filename)
    fp = os.path.join(datasets_dir, safe_name)
    if not os.path.isfile(fp):
        return render_template('DatasetEditor.html', error='Arquivo não encontrado.', filename=safe_name, size_h=None, rows=None)
    try:
        stat = os.stat(fp)
        size_h = _human_readable_size(stat.st_size)
        ext = os.path.splitext(safe_name)[1].lower()
        rows = _count_xlsx_rows(fp) if ext == '.xlsx' else _count_csv_rows(fp)
    except Exception:
        size_h, rows = None, None
    return render_template('DatasetEditor.html', error=None, filename=safe_name, size_h=size_h, rows=rows)

# -----------------------------
# Dataset Management Endpoints
# -----------------------------

def _human_readable_size(num_bytes):
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    size = float(num_bytes)
    for unit in units:
        if size < 1024.0:
            return f"{size:.2f} {unit}"
        size /= 1024.0
    return f"{size:.2f} PB"

def _count_csv_rows(file_path):
    try:
        import pandas as pd
        df = pd.read_csv(file_path)
        # pandas trata cabeçalho automaticamente (header='infer')
        return int(df.shape[0])
    except Exception:
        try:
            # Fallback manual com detecção simples de cabeçalho
            count = 0
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for _ in f:
                    count += 1
            first_line = ''
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_line = f.readline()
            has_header = (',' in first_line) and any(c.isalpha() for c in first_line)
            return max(0, count - (1 if has_header else 0))
        except Exception:
            return None

def _count_xlsx_rows(file_path):
    """Conta linhas de uma planilha XLSX.
    Tenta primeiro via pandas (header tratado automaticamente), depois via openpyxl,
    e por fim usa leitura zip/XML como último fallback sem dependências externas.
    """
    try:
        import pandas as pd
        df = pd.read_excel(file_path)  # usa engine disponível (openpyxl normalmente)
        return int(df.shape[0])
    except Exception:
        pass
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        max_row = ws.max_row or 0
        # Header se a primeira linha tem ao menos uma string
        first_row_cells = [cell.value for cell in ws[1]] if max_row >= 1 else []
        has_header = any(isinstance(v, str) for v in first_row_cells if v is not None)
        wb.close()
        return max(0, max_row - (1 if has_header else 0))
    except Exception:
        pass
    # Fallback leve: lê o XML do primeiro worksheet via zipfile
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(file_path) as zf:
            # usa a primeira planilha encontrada
            sheet_names = [n for n in zf.namelist() if n.startswith('xl/worksheets/') and n.endswith('.xml')]
            if not sheet_names:
                return None
            with zf.open(sheet_names[0]) as xmlf:
                rows = 0
                # iterparse para memória constante
                for event, elem in ET.iterparse(xmlf, events=('end',)):
                    # tags possuem namespace, então usamos endswith
                    if elem.tag.endswith('row'):
                        rows += 1
                # não conseguimos detectar cabeçalho de forma robusta aqui; subtrai 1 conservadoramente
                return max(0, rows - 1)
    except Exception:
        return None

@app.route('/api/datasets', methods=['GET'])
def api_list_datasets():
    items = []
    try:
        for filename in os.listdir(datasets_dir):
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ('.xlsx', '.csv'):
                continue
            fp = os.path.join(datasets_dir, filename)
            try:
                stat = os.stat(fp)
                size_bytes = stat.st_size
                rows = _count_xlsx_rows(fp) if ext == '.xlsx' else _count_csv_rows(fp)
                items.append({
                    'name': filename,
                    'display_name': os.path.splitext(filename)[0],
                    'size_bytes': size_bytes,
                    'size': _human_readable_size(size_bytes),
                    'rows': rows,
                    'created_at': stat.st_ctime,
                    'updated_at': stat.st_mtime,
                    'download_url': f"/api/datasets/download/{filename}",
                    'edit_url': f"/DatasetEditor.html?file={filename}",
                    'delete_url': f"/api/datasets/delete/{filename}"
                })
            except Exception:
                # Skip files that cause issues
                continue
        # Sort by updated_at desc
        items.sort(key=lambda x: x.get('updated_at') or 0, reverse=True)
        return jsonify(items)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/datasets/upload', methods=['POST'])
def api_upload_dataset():
    if 'file' not in request.files:
        return jsonify({'error': 'Arquivo não enviado (campo "file").'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nome de arquivo vazio.'}), 400
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Apenas arquivos .xlsx são suportados.'}), 400
    # Permitir que o cliente envie um nome amigável do dataset
    dataset_name = (request.form.get('dataset_name') or '').strip()
    if dataset_name:
        # Sanitiza e garante extensão .csv
        safe_name = secure_filename(dataset_name)
        filename = safe_name if safe_name.lower().endswith('.xlsx') else f"{safe_name}.xlsx"
    else:
        filename = secure_filename(file.filename)
    save_path = os.path.join(datasets_dir, filename)
    # Se existir, criar um nome único simples
    base, ext = os.path.splitext(filename)
    i = 1
    while os.path.exists(save_path):
        filename = f"{base}_{i}{ext}"
        save_path = os.path.join(datasets_dir, filename)
        i += 1
    file.save(save_path)
    # Build metadata
    stat = os.stat(save_path)
    meta = {
        'name': filename,
        'display_name': os.path.splitext(filename)[0],
        'size_bytes': stat.st_size,
        'size': _human_readable_size(stat.st_size),
        'rows': _count_xlsx_rows(save_path),
        'created_at': stat.st_ctime,
        'updated_at': stat.st_mtime,
        'download_url': f"/api/datasets/download/{filename}",
        'edit_url': f"/DatasetEditor.html?file={filename}"
    }
    return jsonify(meta), 201

@app.route('/api/datasets/download/<path:filename>', methods=['GET'])
def api_download_dataset(filename):
    try:
        return send_from_directory(datasets_dir, filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/datasets/delete/<path:filename>', methods=['DELETE'])
def api_delete_dataset(filename):
    try:
        base_dir = os.path.abspath(datasets_dir)
        target_path = os.path.abspath(os.path.join(base_dir, filename))
        if not target_path.startswith(base_dir):
            return jsonify({'error': 'Caminho inválido.'}), 400
        if not os.path.exists(target_path):
            return jsonify({'error': 'Arquivo não encontrado.'}), 404
        os.remove(target_path)
        return jsonify({'deleted': filename}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/datasets/rename', methods=['POST'])
def api_rename_dataset():
    try:
        data = request.get_json(silent=True, force=True) or {}
        old_name = (data.get('old_name') or '').strip()
        new_name = (data.get('new_name') or '').strip()
        if not old_name or not new_name:
            return jsonify({'error': 'Parâmetros inválidos.'}), 400
        old_base = os.path.basename(old_name)
        new_base = os.path.basename(new_name)
        old_path = os.path.join(datasets_dir, old_base)
        # Preservar extensão se o novo nome vier sem extensão
        if os.path.splitext(new_base)[1] == '':
            ext = os.path.splitext(old_base)[1]
            new_base = os.path.splitext(new_base)[0] + ext
        new_path = os.path.join(datasets_dir, new_base)
        if not os.path.isfile(old_path):
            return jsonify({'error': 'Arquivo não encontrado.'}), 404
        if os.path.exists(new_path):
            return jsonify({'error': 'Já existe um arquivo com esse nome.'}), 409
        os.rename(old_path, new_path)
        return jsonify({'message': 'Renomeado com sucesso.', 'new_name': new_base}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/models', methods=['GET'])
def get_models_api():
    # Endpoint API para obter a lista de modelos com metadados e filename
    models_raw = registry.listar_modelos()
    # Group by model type name for UI
    grouped = {}
    for m in models_raw:
        raw_tipo = m.get('nome_modelo') or m.get('tipo') or 'Desconhecido'
        # Normalize type names to match frontend expectations
        tipo_map = {
            'DecisionTreeClassifier': 'Decision Tree',
            'RandomForestClassifier': 'Random Forest'
        }
        tipo = tipo_map.get(raw_tipo, raw_tipo)
        grouped.setdefault(tipo, []).append({
            'filename': m.get('filename'),
            'nome_teste': m.get('nome_teste'),
            'nome_modelo': tipo,
            'params': {k: v for k, v in m.items() if k not in ['filename','nome_teste','nome_modelo','timestamp','id']}
        })
    return jsonify(grouped)

@app.route('/api/train/start', methods=['POST'])
def api_train_start():
    payload = request.json
    job_id = str(int(time.time() * 1000))
    # Safely parse epochs from payload; handle null/NaN/missing
    tp = payload.get('training_params', {})
    raw_epochs = tp.get('epochs')
    try:
        total_epochs = int(raw_epochs) if raw_epochs is not None else 50
        if total_epochs <= 0:
            total_epochs = 50
    except Exception:
        total_epochs = 50
    # Adjust total epochs for Random Forest by multiplying n_estimators
    try:
        if payload.get('model', {}).get('type') == 'rf':
            cfg = _load_config_by_filename(payload['model']['config'])
            n_est = cfg.get('rf_n_estimators', 10)
            if isinstance(n_est, str):
                n_est = 10 if n_est.strip().lower() in ('none', 'null', '') else int(n_est)
            if not isinstance(n_est, int) or n_est <= 0:
                n_est = 10
            total_epochs = total_epochs * n_est
    except Exception:
        pass
    with TRAIN_JOBS_LOCK:
        TRAIN_JOBS[job_id] = {
            'status': 'queued',
            'epoch': 0,
            'total_epochs': total_epochs,
            'loss': None,
            'accuracy': None,
            'eta_seconds': None,
            'cancel_requested': False
        }
    # Resolver dataset: pode ser enviado como objeto (features/labels) ou como arquivo registrado
    dataset_payload = payload.get('dataset')
    dataset_file = payload.get('dataset_file')

    def _load_dataset_from_file(filename):
        try:
            base = os.path.basename(filename)
            base_dir = os.path.abspath(datasets_dir)
            target_path = os.path.abspath(os.path.join(base_dir, base))
            if not target_path.startswith(base_dir):
                raise ValueError('Caminho inválido para dataset.')
            if not os.path.isfile(target_path):
                raise ValueError('Arquivo de dataset não encontrado.')
            ext = os.path.splitext(base)[1].lower()
            # Leitura com pandas
            import pandas as pd
            if ext == '.xlsx':
                df = pd.read_excel(target_path)
            elif ext == '.csv':
                df = pd.read_csv(target_path)
            else:
                raise ValueError('Formato de dataset não suportado. Use .xlsx ou .csv')

            # Normalizar nomes e tentar identificar colunas de forma flexível
            original_cols = list(df.columns)
            norm_map = {}
            for c in original_cols:
                key = str(c).strip().lower()
                key = key.replace(' ', '')
                key = key.replace('_', '')
                # padronizar x_1 -> x1, x-1 -> x1 etc.
                key = key.replace('-', '')
                norm_map[key] = c

            # Candidatos para a coluna de rótulo
            label_candidates = ['d', 'diagnostico', 'diagnóstico', 'label', 'class', 'target', 'y']
            label_col = None
            for k in label_candidates:
                if k in norm_map:
                    label_col = norm_map[k]
                    break
            # Se não achar, assumir a última coluna como label
            if label_col is None:
                label_col = original_cols[-1]

            # Candidatos para features (x1,x2,x3) se existirem
            x_cols = []
            for k in ['x1','x2','x3','x4','x5']:
                if k in norm_map and norm_map[k] != label_col:
                    x_cols.append(norm_map[k])

            # Se não houver colunas x*, pegar todas exceto label e manter as numéricas
            if not x_cols:
                feature_candidates = [c for c in original_cols if c != label_col]
                # Tentar converter para numérico para filtrar
                df_num = df[feature_candidates].apply(pd.to_numeric, errors='coerce')
                # Selecionar colunas com pelo menos algum valor numérico válido
                valid_feature_cols = [c for c in feature_candidates if df_num[c].notna().any()]
                # Se ainda vazio, erro
                if not valid_feature_cols:
                    raise ValueError('Não foram encontradas colunas numéricas para features.')
                x_cols = valid_feature_cols

            # Construir dataframe numérico final e limpar NaNs
            df_features = df[x_cols].apply(pd.to_numeric, errors='coerce')
            df_labels_raw = df[label_col]
            mask_valid = df_features.notna().all(axis=1) & df_labels_raw.notna()
            df_features = df_features[mask_valid]
            df_labels_raw = df_labels_raw[mask_valid]

            if len(df_features) == 0:
                raise ValueError('Dataset sem linhas válidas após limpeza.')

            # Converter labels em {-1,1}
            def _to_label(v):
                if isinstance(v, bool):
                    return 1 if v else -1
                try:
                    f = float(v)
                    # Mapear 0/1, -1/1 etc.
                    if f in (0.0, 0):
                        return -1
                    return 1 if f > 0 else -1
                except Exception:
                    s = str(v).strip().lower()
                    neg = {'-1','neg','negativo','no','não','nao','false','f'}
                    pos = {'1','pos','positivo','yes','sim','true','t'}
                    if s in neg:
                        return -1
                    if s in pos:
                        return 1
                    # fallback: qualquer outro texto vira 1
                    return 1

            labels = [ _to_label(v) for v in df_labels_raw.tolist() ]
            features = df_features.values.tolist()

            return { 'features': features, 'labels': labels }
        except Exception as e:
            raise e

    if dataset_file and not dataset_payload:
        try:
            dataset_payload = _load_dataset_from_file(dataset_file)
        except Exception as e:
            with TRAIN_JOBS_LOCK:
                TRAIN_JOBS[job_id].update({ 'status': 'error', 'error': f'Dataset inválido: {str(e)}' })
            return jsonify({'error': f'Dataset inválido: {str(e)}'}), 400

    if not dataset_payload:
        with TRAIN_JOBS_LOCK:
            TRAIN_JOBS[job_id].update({ 'status': 'error', 'error': 'Dataset não fornecido.' })
        return jsonify({'error': 'Dataset não fornecido.'}), 400

    _start_training_thread(job_id, payload['model']['type'], payload['model']['config'], dataset_payload, payload.get('training_params', {}))
    return jsonify({'job_id': job_id})

@app.route('/api/train/status/<job_id>', methods=['GET'])
def api_train_status(job_id):
    with TRAIN_JOBS_LOCK:
        job = TRAIN_JOBS.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    # Remove non-serializable objects from the response
    sanitized = dict(job)
    if 'model_ref' in sanitized:
        sanitized.pop('model_ref')
    return jsonify(sanitized)

@app.route('/api/train/cancel/<job_id>', methods=['POST'])
def api_train_cancel(job_id):
    with TRAIN_JOBS_LOCK:
        job = TRAIN_JOBS.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        if job.get('status') in ['completed', 'error', 'canceled']:
            return jsonify({'status': job.get('status')}), 200
        job['cancel_requested'] = True
        TRAIN_JOBS[job_id] = job
    return jsonify({'status': 'cancelling'})

@app.route('/api/train/result/<job_id>', methods=['GET'])
def api_train_result(job_id):
    with TRAIN_JOBS_LOCK:
        job = TRAIN_JOBS.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'completed':
        return jsonify({'error': 'Job not completed yet'}), 400
    return jsonify(job.get('results'))

@app.route('/api/test/<job_id>', methods=['POST'])
def api_test_model(job_id):
    # Test a trained model with provided dataset, return predictions and metrics
    payload = request.json
    features = np.array(payload['features'])
    labels = np.array(payload.get('labels')) if payload.get('labels') is not None else None
    with TRAIN_JOBS_LOCK:
        job = TRAIN_JOBS.get(job_id)
    if not job or job.get('status') != 'completed':
        return jsonify({'error': 'Model not ready'}), 400
    model = job.get('model_ref')
    if model is None:
        return jsonify({'error': 'Model reference unavailable'}), 500
    # Predict
    y_pred = model.predict(features).tolist()
    response = {
        'predictions': y_pred
    }
    # If labels provided, compute metrics
    if labels is not None:
        # Convert to numpy and metrics consistent with trainingModels evaluate
        from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
        yp = np.array(y_pred)
        acc = float(accuracy_score(labels, yp))
        prec = float(precision_score(labels, yp, zero_division=0))
        rec = float(recall_score(labels, yp, zero_division=0))
        f1 = float(f1_score(labels, yp, zero_division=0))
        cm = confusion_matrix(labels, yp, labels=[-1, 1]).tolist()

        # Try to provide raw probabilities if the model supports it
        proba = None
        try:
            if hasattr(model, 'predict_proba'):
                proba = model.predict_proba(features).tolist()
        except Exception:
            proba = None

        response.update({
            'accuracy': acc,
            'precision': prec,
            'recall': rec,
            'f1_score': f1,
            'confusion_matrix': cm,
            'probabilities': proba
        })
    return jsonify(response)

@app.route('/cadastrar_modelo', methods=['POST'])
def cadastrar_modelo():
    nome_teste = request.form['id_teste']
    nome_modelo = request.form['modelo_ml_utilizado']

    if nome_modelo == 'Regressão Logística':
        registry.cadastrar_logistic(
            nome_teste,
            request.form.get('estrategia_preprocessamento'),
            request.form.get('estrategia_validacao'),
            request.form.get('logreg_penalty'),
            request.form.get('logreg_C'),
            request.form.get('logreg_solver'),
            request.form.get('logreg_class_weight'),
            request.form.get('logreg_max_iter')
        )
    elif nome_modelo == 'KNN':
        registry.cadastrar_knn(
            nome_teste,
            request.form.get('estrategia_preprocessamento'),
            request.form.get('estrategia_validacao'),
            request.form.get('knn_n_neighbors'),
            request.form.get('knn_weights'),
            request.form.get('knn_metric'),
            request.form.get('knn_p')
        )
    elif nome_modelo == 'SVM':
        registry.cadastrar_svm( 
            nome_teste,
            request.form.get('estrategia_preprocessamento'),
            request.form.get('estrategia_validacao'),
            request.form.get('svm_C'),
            request.form.get('svm_kernel'),
            request.form.get('svm_gamma'),
            request.form.get('svm_degree'),
            request.form.get('svm_class_weight'),
            request.form.get('svm_probability')
        )
    elif nome_modelo == 'DecisionTreeClassifier':
        registry.cadastrar_tree(
            nome_teste,
            request.form.get('estrategia_preprocessamento'),
            request.form.get('estrategia_validacao'),
            request.form.get('tree_criterion'),
            request.form.get('tree_max_depth'),
            request.form.get('tree_min_samples_split'),
            request.form.get('tree_min_samples_leaf'),
            request.form.get('tree_class_weight')
        )
    elif nome_modelo == 'RandomForestClassifier':
        registry.cadastrar_rf(
            nome_teste,
            request.form.get('estrategia_preprocessamento'),
            request.form.get('estrategia_validacao'),
            request.form.get('rf_n_estimators'),
            request.form.get('rf_max_depth'),
            request.form.get('rf_criterion'),
            request.form.get('rf_min_samples_split'),
            request.form.get('rf_min_samples_leaf'),
            request.form.get('rf_class_weight'),
            request.form.get('rf_oob_score')
        )
    return redirect('/Models.html')



@app.route('/api/train-compare', methods=['POST'])
def train_compare():
    data = request.json
    
    # Extract dataset
    features = np.array(data['dataset']['features'])
    labels = np.array(data['dataset']['labels'])
    
    # Split dataset
    test_size = data['training_params']['test_split']
    X_train, X_test, y_train, y_test = train_test_split(
        features, labels, test_size=test_size, random_state=42
    )
    
    # Train and evaluate models
    results = {
        'model1': train_and_evaluate_model(data['model1'], X_train, X_test, y_train, y_test, 
                                          data['training_params']),
        'model2': train_and_evaluate_model(data['model2'], X_train, X_test, y_train, y_test,
                                          data['training_params'])
    }
    
    return jsonify(results)

def train_and_evaluate_model(model_info, X_train, X_test, y_train, y_test, training_params):
    model_type = model_info['type']
    config_file = model_info['config']
    
    # Load model configuration
    model_config = load_model_config(model_type, config_file)
    
    # Use the appropriate training function based on model type
    if model_type == 'logistic':
        from trainingModels.LogisticRegression import train_logistic
        results, _ = train_logistic(model_config, X_train, X_test, y_train, y_test, training_params)
    elif model_type == 'knn':
        from trainingModels.KNN import train_knn
        results, _ = train_knn(model_config, X_train, X_test, y_train, y_test, training_params)
    elif model_type == 'svm':
        from trainingModels.SVM import train_svm
        results, _ = train_svm(model_config, X_train, X_test, y_train, y_test, training_params)
    elif model_type == 'tree':
        from trainingModels.DecisionTree import train_tree
        results, _ = train_tree(model_config, X_train, X_test, y_train, y_test, training_params)
    elif model_type == 'rf':
        from trainingModels.RandomForest import train_rf
        results, _ = train_rf(model_config, X_train, X_test, y_train, y_test, training_params)
    else:
        raise ValueError(f"Unsupported model type: {model_type}")
    
    # Add model name to results if not already present
    if 'name' not in results:
        results['name'] = config_file.replace('.json', '')
    
    return results

def load_model_config(model_type, config_file):
    model_type_map = {
        'logistic': 'Regressão Logística',
        'knn': 'KNN',
        'svm': 'SVM',
        'tree': 'Decision Tree',
        'rf': 'Random Forest'
    }
    
    model_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models_json')
    config_path = os.path.join(model_dir, config_file)
    
    with open(config_path, 'r') as f:
        return json.load(f)
    
            

if __name__ == "__main__":
    app.run(debug=True)
